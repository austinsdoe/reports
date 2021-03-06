import os

from DateTime import DateTime
from bika.lims import api
from bika.lims.browser import BrowserView
from bika.lims.browser.reports.selection_macros import SelectionMacrosView
from bika.lims.catalog import CATALOG_ANALYSIS_LISTING
from bika.lims.catalog import CATALOG_ANALYSIS_REQUEST_LISTING
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from plone.app.layout.globals.interfaces import IViewView
from plone.memoize import view as viewcache
from zope.interface import implements


def save_in_memory_and_return(wb):
    virtual_workbook = save_virtual_workbook(wb)
    return {'report_title': "Saving report",
            'is_excel': True,
            'report_data': virtual_workbook}


class Report(BrowserView):
    implements(IViewView)

    def __init__(self, context, request, report=None):
        BrowserView.__init__(self, context, request)
        self.report = report
        self.selection_macros = SelectionMacrosView(self.context, self.request)
        self.cells = dict()

    def __call__(self):
        year = self.request.form.get('year', DateTime().year())
        quarter = self.request.form.get('quarter', "Q1")
        category_uid = self.request.form.get('CategoryUID', None)
        if not category_uid:
            return

        date_from = "{}-01-01"
        date_to = "{}-03-31"
        if quarter == "Q1":
            date_from = "{}-10-01"
            date_to = "{}-12-31"
        elif quarter == "Q3":
            date_from = "{}-04-01"
            date_to = "{}-06-30"
        elif quarter == "Q4":
            date_from = "{}-07-01"
            date_to = "{}-09-30"

        date_from = date_from.format(year)
        date_to = date_to.format(year)
        date_from = api.to_date(date_from, DateTime())
        date_to = api.to_date(date_to, DateTime())

        query = {
            'getCategoryUID': category_uid,
            'getDatePublished': {'query': [date_from, date_to],
                                 'range': 'min:max'},
            'review_state': ['verified', 'published'],
            'cancellation_state': 'active',}
        self.cells = dict()
        catalog = api.get_tool(CATALOG_ANALYSIS_LISTING)
        for analysis_brain in catalog(query):
            result = self.to_float(analysis_brain.getResult)
            if not result:
                continue
            patient_brain = self.get_patient_brain(analysis_brain)
            if not patient_brain:
                continue
            self.fill_reported(result)
            self.fill_results_by_sex(result, patient_brain)
            self.fill_results_by_age(result, patient_brain)
            self.fill_results_by_pregnancy(result, analysis_brain)

        # Build report
        this_dir = os.path.dirname(os.path.abspath(__file__))
        templates_dir = os.path.join(this_dir, 'excel_files')
        wb = load_workbook(
            templates_dir + '/ViralLoadQuarterlyMonitoringToolv1.xlsx')
        # grab the active worksheet
        ws = wb.get_sheet_by_name('NMRL')
        for cell_id, value in self.cells.items():
            ws[cell_id] = str(value)

        # Leave header cells empty for now
        laboratory = self.context.bika_setup.laboratory
        address = laboratory.getPhysicalAddress()
        user = api.get_current_user()
        ws["I3"] = address and address.get('state', "") or ""
        ws["L3"] = address and address.get('city', "") or ""
        ws["D3"] = address and address.get('country', "") or ""
        ws["E5"] = laboratory.Title()
        ws["E7"] = laboratory.getTaxNumber()
        ws["I7"] = user.getProperty('fullname')
        ws["L7"] = user.getProperty('email')
        ws["H9"] = quarter

        # Save the file in memory
        return save_in_memory_and_return(wb)

    def get_column_result(self, result):
        if result < 1000:
            return "H"
        return "I"

    def fill_reported(self, result):
        col = self.get_column_result(result)
        cell_id = "{}15".format(col)
        self.add_count_cell(cell_id)
        self.add_count_cell("H12")

    def fill_results_by_sex(self, result, patient_brain):
        gender = patient_brain.getGender
        row = 0
        if gender == 'male':
            row = 17
        elif gender == 'female':
            row = 18
        else:
            return

        col = self.get_column_result(result)
        cell_id = "{}{}".format(col, str(row))
        self.add_count_cell(cell_id)
        # Total:
        cell_id = "{}19".format(col)
        self.add_count_cell(cell_id)

    def fill_results_by_age(self, result, patient_brain):
        age_splitted = patient_brain.getAgeSplittedStr
        if not age_splitted:
            return

        age = 0
        if age_splitted.find("y") > 0:
            age = self.to_float(age_splitted[0:age_splitted.find("y")])
            if not age:
                return

        col = self.get_column_result(result)
        row = 0
        row_subtotal = 0
        if age < 1:
            row = 21
            row_subtotal = 24
        elif age <= 9:
            row = 22
            row_subtotal = 24
        elif age <= 14:
            row = 23
            row_subtotal = 24
        elif age <= 19:
            row = 25
            row_subtotal = 27
        elif age <= 24:
            row = 26
            row_subtotal = 27
        elif age >= 25:
            row = 28
        if not row:
            return

        cell_id = "{}{}".format(col, str(row))
        self.add_count_cell(cell_id)

        if row_subtotal:
            cell_subtotal_id = "{}{}".format(col, str(row_subtotal))
            self.add_count_cell(cell_subtotal_id)

        cell_total_id = "{}29".format(col)
        self.add_count_cell(cell_total_id)

    def fill_results_by_pregnancy(self, result, analysis):
        ar_uid = analysis.getParentUID
        if not ar_uid:
            return
        ar = self.get_brain(ar_uid, CATALOG_ANALYSIS_REQUEST_LISTING)
        if not ar:
            return
        batch_uid = ar.getBatchUID
        if not batch_uid:
            return
        batch = self.get_object(batch_uid)
        if not batch:
            return None

        col = self.get_column_result(result)
        try:
            pregnant = batch.Schema().getField('VLPregnant').get(batch)
            if pregnant:
                cell_id = "{}30".format(col)
                self.add_count_cell(cell_id)
            breast = batch.Schema().getField('VLBreastFeeding').get(batch)
            if breast:
                cell_id = "{}31".format(col)
                self.add_count_cell(cell_id)
        except:
            pass

    def to_float(self, value):
        try:
            return float(value)
        except:
            return None

    def add_count_cell(self, id_cell, count=1):
        current = self.cells.get(id_cell, 0)
        self.cells[id_cell] = current + count

    @viewcache.memoize
    def get_brain(self, uid, catalog):
        cat = api.get_tool(catalog)
        brain = cat(UID=uid)
        if not brain or len(brain) == 0:
            return None
        return brain[0]

    @viewcache.memoize
    def get_object(self, brain_or_object_or_uid):
        """Get the full content object. Returns None if the param passed in is
        not a valid, not a valid object or not found

        :param brain_or_object_or_uid: UID/Catalog brain/content object
        :returns: content object
        """
        if api.is_uid(brain_or_object_or_uid):
            return api.get_object_by_uid(brain_or_object_or_uid, default=None)
        if api.is_object(brain_or_object_or_uid):
            return api.get_object(brain_or_object_or_uid)
        return None

    def get_patient_brain(self, analysis_brain):
        ar_uid = analysis_brain.getParentUID
        if not ar_uid:
            return None
        ar = self.get_brain(ar_uid, CATALOG_ANALYSIS_REQUEST_LISTING)
        if not ar:
            return None
        patient_uid = ar.getPatientUID
        if not patient_uid:
            return None
        patient = self.get_brain(patient_uid, 'bikahealth_catalog_patient_listing')
        return patient
